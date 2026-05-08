import time
import requests
import json
import os
import io
import sys
import traceback
from datetime import datetime

from github import Github, GithubException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook

def upload_to_github(data, filename="data_tagihan_spk.json"):
    token = os.environ.get('GITHUB_TOKEN')
    if not token:
        print("[ERROR] GITHUB_TOKEN tidak ditemukan!")
        return False

    try:
        g = Github(token)
        repo = g.get_repo("ipanrifan-create/TAGIHAN-SPK")
        file_content = json.dumps(data, indent=4, ensure_ascii=False)

        try:
            contents = repo.get_contents(filename)
            repo.update_file(
                path=contents.path,
                message=f"Update data SPK otomatis - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                content=file_content,
                sha=contents.sha
            )
            print(f"[OK] File {filename} berhasil diperbarui di GitHub.")
        except GithubException as e:
            if e.status == 404:
                repo.create_file(
                    path=filename,
                    message="Initial commit data SPK",
                    content=file_content
                )
                print(f"[OK] File {filename} berhasil dibuat di GitHub.")
            else:
                print(f"[ERROR] GitHub API Error: {e}")
                return False
                
        return True
    except Exception as e:
        print(f"[ERROR] Gagal upload ke GitHub: {e}")
        return False

def create_chrome_driver():
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    # Inisialisasi langsung (kompatibel dengan setup-chromedriver di Actions)
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    return driver

def jalankan_bot():
    print("=" * 50)
    print("=== [START] Operasi Bot Tagihan SPK ===")
    print("=" * 50)

    email = os.environ.get('EMAIL_SCM')
    password = os.environ.get('PASS_SCM')

    if not email or not password:
        print("[ERROR] Credentials tidak ditemukan di Environment Variables!")
        return False

    driver = create_chrome_driver()
    wait = WebDriverWait(driver, 20)

    try:
        # ===== STEP 1: Login =====
        print("[STEP 1] Login ke SCM Nusadaya...")
        driver.get("https://scm.nusadaya.net/login")
        
        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @placeholder='Email atau NIP']")))
        email_input.clear()
        email_input.send_keys(email)
        
        password_input = driver.find_element(By.XPATH, "//input[@type='password']")
        password_input.clear()
        password_input.send_keys(password)
        
        driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]").click()
        time.sleep(5)
        print(f"[STEP 1] Login selesai. URL saat ini: {driver.current_url}")

        # ===== STEP 2: Download Data SPK =====
        print("[STEP 2] Proses Download Data SPK...")
        session_cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        
        tahun_sekarang = datetime.now().year
        download_url = f"https://scm.nusadaya.net/surat-perintah-kerja/export?bidang=01&tahun={tahun_sekarang}"
        
        headers = {
            'Referer': 'https://scm.nusadaya.net/surat-perintah-kerja',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        response_dl = requests.get(download_url, cookies=session_cookies, headers=headers, timeout=60)

        if response_dl.status_code == 200:
            print("[STEP 2] Download Excel SPK Berhasil.")
            
            # ===== STEP 3: Konversi Excel ke JSON & Upload =====
            print("[STEP 3] Memproses Excel ke JSON...")
            wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=True)
            all_data = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append([cell if cell is not None else "" for cell in row])
                all_data[sheet_name] = rows

            success = upload_to_github(all_data)
            if success:
                print("[FINISH] Bot berhasil memperbarui data di TAGIHAN-SPK.")
                return True
            else:
                return False
        else:
            print(f"[ERROR] Gagal download. Status code: {response_dl.status_code}")
            return False

    except Exception as e:
        print(f"[FATAL ERROR] {e}")
        traceback.print_exc()
        
        # Otomatis screenshot jika error untuk memudahkan debugging
        try:
            driver.save_screenshot("error_screenshot.png")
            print("[DEBUG] Screenshot error disimpan sebagai error_screenshot.png")
        except:
            pass
            
        return False
    finally:
        print("[INFO] Menutup browser...")
        driver.quit()

if __name__ == "__main__":
    result = jalankan_bot()
    sys.exit(0 if result else 1)
